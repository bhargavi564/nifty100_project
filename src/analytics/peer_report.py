import os
import sqlite3
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Resolve workspace paths
BASE_DIR = Path(__file__).resolve().parents[2]
DB_PATH = BASE_DIR / "data" / "nifty100.db"
OUTPUT_FILE = BASE_DIR / "output" / "peer_comparison.xlsx"


def generate_peer_comparison_report(db_path=DB_PATH, output_file=OUTPUT_FILE):
    print("[INFO] Starting Day 20 Peer Comparison Excel Report Generation...")
    
    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if not os.path.exists(db_path):
        print(f"[CRITICAL] Database file missing at: {db_path}")
        return "Database missing"

    conn = sqlite3.connect(db_path)
    
    # 1. Pull raw values and calculated percentiles
    query = """
        SELECT company_id, peer_group_name, metric, value, percentile_rank, year 
        FROM peer_percentiles
    """
    try:
        df = pd.read_sql_query(query, conn)
    except Exception as e:
        print(f"[ERROR] Database query failed: {e}")
        conn.close()
        return "Database query failed"
    finally:
        conn.close()

    if df.empty:
        print("[WARNING] No data available to generate report.")
        return "No data found"

    # Add mock names if company_name column is missing in intermediate stages
    df["company_name"] = df["company_id"] + " Ltd."

    # Identify distinct peer groups
    peer_groups = [g for g in df["peer_group_name"].unique() if pd.notna(g)]

    # We'll store pre-calculated medians here to write them out during the openpyxl pass
    medians_dict = {}

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for group in peer_groups:
            group_df = df[df["peer_group_name"] == group].copy()
            
            # 2. Wide pivoting for 20 metrics and their corresponding percentile ranks
            pivoted_vals = group_df.pivot_table(
                index=["company_id", "company_name"], 
                columns="metric", 
                values="value", 
                aggfunc="first"
            ).reset_index()

            pivoted_ranks = group_df.pivot_table(
                index=["company_id", "company_name"], 
                columns="metric", 
                values="percentile_rank", 
                aggfunc="first"
            ).reset_index()

            # Merge columns and add suffix to percentile rank headers
            pivoted_ranks = pivoted_ranks.rename(
                columns={col: f"{col}_pct_rank" for col in pivoted_ranks.columns if col not in ["company_id", "company_name"]}
            )
            
            sheet_data = pd.merge(pivoted_vals, pivoted_ranks, on=["company_id", "company_name"], how="outer")
            
            # Sort columns so the percentile rank sits right next to its raw metric
            sorted_cols = ["company_id", "company_name"]
            raw_metrics = [c for c in pivoted_vals.columns if c not in ["company_id", "company_name"]]
            
            for metric in raw_metrics:
                sorted_cols.append(metric)
                pct_col = f"{metric}_pct_rank"
                if pct_col in sheet_data.columns:
                    sorted_cols.append(pct_col)

            sheet_data = sheet_data[sorted_cols]
            
            # Pre-calculate pandas numeric medians for the bottom summary row
            medians = {}
            for col in sheet_data.columns:
                if col not in ["company_id", "company_name"]:
                    medians[col] = pd.to_numeric(sheet_data[col], errors="coerce").median()
            
            sheet_name = str(group)[:31]
            medians_dict[sheet_name] = medians

            # Write pivoted sheet structure to Excel
            sheet_data.to_excel(writer, sheet_name=sheet_name, index=False)

    # 4. Open with openpyxl to apply styling and write pre-calculated flat medians
    wb = load_workbook(output_path)
    
    # Define color fills
    green_fill = PatternFill(fill_type="solid", start_color="C6EFCE")  # Green (>= 75th pct)
    yellow_fill = PatternFill(fill_type="solid", start_color="FFEB9C") # Yellow (25th to 75th pct)
    red_fill = PatternFill(fill_type="solid", start_color="FFC7CE")    # Red (<= 25th pct)
    gold_fill = PatternFill(fill_type="solid", start_color="FFD700")   # Gold/Amber for Benchmark
    gray_fill = PatternFill(fill_type="solid", start_color="F2F2F2")   # Gray for Summary Medians

    bold_font = Font(bold=True)
    border_thin_top_double_bottom = Border(
        top=Side(style='thin'), 
        bottom=Side(style='double')
    )

    for ws in wb.worksheets:
        max_row = ws.max_row
        max_col = ws.max_column
        headers = [ws.cell(1, col).value for col in range(1, max_col + 1)]
        sheet_name = ws.title

        # Determine which row is the group benchmark company (We pick the 1st row as our sample benchmark)
        benchmark_row_index = 2 if max_row >= 2 else None

        # Apply formatting and styles cell by cell
        for row in range(2, max_row + 1):
            is_benchmark = (row == benchmark_row_index)
            
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                header = headers[col - 1]

                # Base highlight: Row background for benchmark company
                if is_benchmark:
                    cell.fill = gold_fill
                
                # Check if this column is a percentile rank column
                if header and "_pct_rank" in header:
                    try:
                        val = float(cell.value) if cell.value is not None else None
                        if val is not None:
                            if val >= 0.75:
                                cell.fill = green_fill
                            elif val <= 0.25:
                                cell.fill = red_fill
                            else:
                                cell.fill = yellow_fill
                    except ValueError:
                        pass

        # 5. Append pre-calculated flat pandas median values to bypass Excel formula rendering errors
        median_row_num = max_row + 1
        ws.cell(row=median_row_num, column=1, value="Group Median").font = bold_font
        ws.cell(row=median_row_num, column=2, value="-").alignment = Alignment(horizontal="center")
        
        group_medians = medians_dict.get(sheet_name, {})

        for col in range(3, max_col + 1):
            header = headers[col - 1]
            median_cell = ws.cell(row=median_row_num, column=col)
            
            # Pull pre-calculated numeric value instead of writing dynamic "=MEDIAN(...)" string
            raw_median_val = group_medians.get(header, np.nan)
            
            # Write a clean, flat number or string representation
            median_cell.value = "" if pd.isna(raw_median_val) else round(float(raw_median_val), 4)
            median_cell.font = bold_font
            median_cell.fill = gray_fill
            median_cell.border = border_thin_top_double_bottom

    wb.save(output_path)
    print(f"[SUCCESS] Exported styled peer comparison report with flat values to: {output_path}")
    return "Report generated successfully"


if __name__ == "__main__":
    generate_peer_comparison_report()