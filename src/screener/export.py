from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def export_results(data, output_file="output/screener_output.xlsx"):
    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # 1. Write DataFrames to Excel sheets
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        if isinstance(data, dict):
            for sheet_name, df in data.items():
                if isinstance(df, pd.DataFrame):
                    df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
        else:
            data.to_excel(writer, sheet_name="Sheet1", index=False)

    # 2. Reload workbook using openpyxl to apply conditional color styling
    wb = load_workbook(output_path)
    green = PatternFill(fill_type="solid", start_color="C6EFCE")  # Soft green
    red = PatternFill(fill_type="solid", start_color="FFC7CE")    # Soft red

    for ws in wb.worksheets:
        # Edge-case safety: check if the sheet actually has rows
        if ws.max_row < 1:
            continue

        headers = [c.value for c in ws[1]]
        if "composite_quality_score" not in headers:
            continue

        score_col = headers.index("composite_quality_score") + 1

        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=score_col)
            try:
                score = float(cell.value)
            except (TypeError, ValueError):
                continue
            
            # Apply styling based on score threshold
            cell.fill = green if score >= 70 else red

    wb.save(output_path)
    return str(output_path)


# =====================================================================
# Execution Demo
# =====================================================================
if __name__ == "__main__":
    print("=" * 60)
    print("Running Excel Export & Styling Demo...")
    print("=" * 60)

    # Generate mock data categories
    high_quality_stocks = pd.DataFrame({
        "ticker": ["TCS", "INFY", "ITC"],
        "composite_quality_score": [98, 92, 88],
        "return_on_equity_pct": [38.2, 29.1, 25.4]
    })

    high_risk_stocks = pd.DataFrame({
        "ticker": ["STOCK_X", "STOCK_Y", "STOCK_Z"],
        "composite_quality_score": [65, 45, 72],
        "return_on_equity_pct": [8.5, 5.0, 12.1]
    })

    # Combine into sheets
    screener_results = {
        "Quality Compounders": high_quality_stocks,
        "Turnaround Candidates": high_risk_stocks
    }

    # Run the export function
    output_path = export_results(screener_results)
    print(f"\n[SUCCESS] Excel sheet generated and styled at: {output_path}")
    print("=" * 60)