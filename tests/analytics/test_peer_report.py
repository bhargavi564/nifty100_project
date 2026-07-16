import os
import sys
import unittest
import shutil
import sqlite3
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

# Resolve workspace paths
BASE_DIR = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(BASE_DIR / "src" / "analytics"))
sys.path.insert(0, str(BASE_DIR))

from peer_report import generate_peer_comparison_report


class TestPeerComparisonReport(unittest.TestCase):

    def setUp(self):
        """Set up isolated mock test folders and SQLite tables."""
        self.test_dir = BASE_DIR / "tests" / "temp_report_outputs"
        self.test_db = self.test_dir / "test_nifty100.db"
        self.test_excel = self.test_dir / "peer_comparison.xlsx"
        self.test_dir.mkdir(parents=True, exist_ok=True)

        self.conn = sqlite3.connect(self.test_db)
        cursor = self.conn.cursor()
        cursor.execute("""
            CREATE TABLE peer_percentiles (
                company_id TEXT,
                peer_group_name TEXT,
                metric TEXT,
                value REAL,
                percentile_rank REAL,
                year INTEGER
            )
        """)

        # Add mock data
        mock_records = [
            ("TCS", "IT", "ROE", 40.0, 0.95, 2026),
            ("TCS", "IT", "D/E", 0.02, 0.90, 2026),
            ("INFY", "IT", "ROE", 30.0, 0.15, 2026),
            ("INFY", "IT", "D/E", 0.00, 0.99, 2026),
        ]
        cursor.executemany("INSERT INTO peer_percentiles VALUES (?, ?, ?, ?, ?, ?)", mock_records)
        self.conn.commit()
        self.conn.close()

    def tearDown(self):
        """Clean up test folders."""
        if self.test_dir.exists():
            shutil.rmtree(self.test_dir)

    def test_report_sheet_and_formatting_generation(self):
        """Should write standard columns, percentiles, color cells, and pre-calculated medians."""
        status = generate_peer_comparison_report(db_path=self.test_db, output_file=self.test_excel)
        self.assertEqual(status, "Report generated successfully")
        self.assertTrue(self.test_excel.exists())

        # Load file and assert structural properties
        wb = load_workbook(self.test_excel)
        self.assertIn("IT", wb.sheetnames)
        
        ws = wb["IT"]
        headers = [cell.value for cell in ws[1]]
        
        # Verify wide pivoting structure (Raw metric immediately followed by percentile rank)
        self.assertIn("company_id", headers)
        self.assertIn("ROE", headers)
        self.assertIn("ROE_pct_rank", headers)

        # Check group median label
        median_row = ws.max_row
        self.assertEqual(ws.cell(row=median_row, column=1).value, "Group Median")
        
        # Verify that the written median matches the pre-calculated numeric value of 0.01 instead of a formula string
        self.assertEqual(float(ws.cell(row=median_row, column=3).value), 0.01)


if __name__ == "__main__":
    unittest.main()